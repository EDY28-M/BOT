
import re
import pandas as pd
from typing import List, Dict, BinaryIO, Any
from io import BytesIO

class ExcelService:
    @staticmethod
    def parse_uploaded_file(file: BinaryIO, filename: str) -> Dict[str, Any]:
        """
        Parsea el archivo subido y separa DNIs válidos (8 dígitos) de inválidos.
        Retorna: {"valid": [...], "invalid": [...]}
        """
        raw_entries = []
        try:
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                df = pd.read_excel(BytesIO(file.read()))
                col = df.columns[0]
                if 'DNI' in df.columns:
                    col = 'DNI'
                raw_entries = df[col].astype(str).tolist()

            elif filename.endswith(".txt") or filename.endswith(".csv"):
                content = file.read().decode("utf-8", errors="ignore")
                raw_entries = content.splitlines()

        except Exception as e:
            raise ValueError(f"Error parseando archivo: {e}")

        valid_dnis = []
        invalid_dnis = []
        seen = set()

        for raw in raw_entries:
            # Limpiar: quitar espacios, manejar "12345678.0" de Excel
            clean = raw.strip().split(".")[0].strip()
            if not clean or clean.lower() == 'nan':
                continue

            # Validar: exactamente 8 dígitos numéricos
            if re.fullmatch(r'\d{8}', clean):
                if clean not in seen:
                    seen.add(clean)
                    valid_dnis.append(clean)
            else:
                invalid_dnis.append(clean)

        return {"valid": valid_dnis, "invalid": invalid_dnis}

    @staticmethod
    def _style_sheet(ws, df, table_name):
        """Aplica estilos a una hoja: header gris, colores por estado, tabla, auto-ancho."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter

        font_normal = Font(name='Aptos Narrow', size=11)
        font_header = Font(name='Aptos Narrow', size=11, bold=True, color='000000')
        fill_header = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF'),
        )
        align = Alignment(vertical='center', wrap_text=False)

        num_rows = len(df)
        num_cols = len(df.columns)
        if num_cols == 0:
            return

        # Header
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = thin_border
            cell.alignment = align

        # Encontrar columna Estado
        estado_col = None
        for ci in range(1, num_cols + 1):
            if ws.cell(row=1, column=ci).value == 'Estado':
                estado_col = ci
                break

        fill_found = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        fill_not_found = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        estados_found = {'FOUND_SUNEDU', 'FOUND_MINEDU'}
        estados_not_found = {'NOT_FOUND', 'ERROR_SUNEDU', 'ERROR_MINEDU'}

        # Datos + colores
        for row_idx in range(2, num_rows + 2):
            row_fill = None
            if estado_col:
                estado_val = str(ws.cell(row=row_idx, column=estado_col).value or '')
                if estado_val in estados_found:
                    row_fill = fill_found
                elif estado_val in estados_not_found:
                    row_fill = fill_not_found

            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_normal
                cell.border = thin_border
                cell.alignment = align
                if row_fill:
                    cell.fill = row_fill

        # Tabla Excel
        if num_rows > 0:
            table_ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
            table = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleLight1",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=False, showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        # Auto-ancho
        for col_idx in range(1, num_cols + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ''))
            for row_idx in range(2, min(num_rows + 2, 102)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    @staticmethod
    def generate_excel(data: List[dict]) -> BytesIO:
        df_all = pd.DataFrame(data)
        df_sunedu = df_all[df_all['Estado'] == 'FOUND_SUNEDU'].reset_index(drop=True) if 'Estado' in df_all.columns else pd.DataFrame()
        df_minedu = df_all[df_all['Estado'] == 'FOUND_MINEDU'].reset_index(drop=True) if 'Estado' in df_all.columns else pd.DataFrame()

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja 1: Todos
            df_all.to_excel(writer, index=False, sheet_name='Todos')
            ExcelService._style_sheet(writer.sheets['Todos'], df_all, 'TablaTodos')

            # Hoja 2: Sunedu
            df_sunedu.to_excel(writer, index=False, sheet_name='Sunedu')
            ExcelService._style_sheet(writer.sheets['Sunedu'], df_sunedu, 'TablaSunedu')

            # Hoja 3: Minedu
            df_minedu.to_excel(writer, index=False, sheet_name='Minedu')
            ExcelService._style_sheet(writer.sheets['Minedu'], df_minedu, 'TablaMinedu')

        output.seek(0)
        return output

